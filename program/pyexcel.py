import openpyxl

# 创建excel
wb=openpyxl.Workbook()
# 激活 worksheet
ws = wb.active #默认
# ws1 = wb.create_sheet("Mysheet") #插入到最后(default)
# ws2 = wb.create_sheet("Mysheet", 0) #插入到最开始的位置
ws.title = "NewTitle" #名称
ws.sheet_properties.tabColor = "000000" #页签颜色
wb.sheetnames #获取所有sheet
ws['A1'] = 123
# 数据可以直接分配到单元格中
ws['A1'] = 42
# 可以附加行，从第一列开始附加
ws.append([1, 2, 3])
# 保存文件
wb.save("file/Python.xlsx")

# 打开excel
# wb=openpyxl.load_workbook('1234.xlsx')